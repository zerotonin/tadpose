# ╔══════════════════════════════════════════════════════════════════╗
# ║  TadPose — metadata_ingest                                       ║
# ║  « spreadsheet in, plate groups and per-video layouts out »      ║
# ╠══════════════════════════════════════════════════════════════════╣
# ║  Turn the lab metadata spreadsheet into the database records     ║
# ║  and per-video plate-layout CSVs the pipeline ingests, with no   ║
# ║  interactive PlateManager / ExperimentManager menus.             ║
# ║                                                                  ║
# ║  Two phases, so a sheet full of typos never silently mislabels   ║
# ║  a genotype:                                                     ║
# ║                                                                  ║
# ║    plan    parse the sheet, match each video to its row, and     ║
# ║            write a reviewable plan CSV (one line per video).     ║
# ║    commit  read the (edited) plan, get-or-create Frog /          ║
# ║            TadpoleGroup / WellType / Investigator, and write     ║
# ║            each video a meta_data_table.csv under its own dir.   ║
# ║                                                                  ║
# ║  Each video is one plate of one (clutch, gene) group across      ║
# ║  its occupied wells; empty wells stay None and drop at ingest.   ║
# ╚══════════════════════════════════════════════════════════════════╝
"""Spreadsheet in, plate groups and per-video layouts out.

Turn the lab METADATA TADPOLES.xlsx into the database records and per-video
plate-layout CSVs the pipeline ingests, replacing the interactive PlateManager
and ExperimentManager menus with a reviewable plan -> commit flow.
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from tadpose import config
from tadpose.database import (
    DatabaseHandler,
    Frog,
    Investigator,
    TadpoleGroup,
    WellType,
)

N_WELLS = 24
DEFAULT_EXPERIMENT_TYPE_ID = 7          # "ND2 31/11/24" Nov-2024 session
DEFAULT_WELL_TYPE_NAME = "medium"       # plain rearing medium, no drug
DEFAULT_WELL_TYPE_DESC = "plain rearing medium, no drug"

# Sheet field labels (row 2 of the two-row header) we care about, matched by
# case-insensitive substring so minor wording drift does not break parsing.
_FIELD_KEYS: dict[str, str] = {
    "key": "key",
    "initials": "initials",
    "date of fertilisation": "fert_date",
    "female tank": "female_tank",
    "female identifier": "female_identifier",
    "background strain": "background_strain",
    "neurod2": "target_gene",            # "eg NeuroD2 L and S"
    "cas9": "cas9",
    "sgrna": "sgrna",                    # "sgRNA rank, dilution"
    "sequence folder": "seq_folder",
    "date of experiment": "exp_date",
    "stage of tadpole": "stage",
    "control arenas": "control_arenas",
    "test arenas": "test_arenas",
    "drug and dose": "drug_dose",
    "notes": "notes",
}

# Dose / drug tokens that mark a per-well gradient plate (not a uniform group).
_GRADIENT_RE = re.compile(r"\b(\d+\s*mm|ptz|4-?ap|vpa)\b", re.IGNORECASE)


# ┌────────────────────────────────────────────────────────────┐
# │ Well-range parsing  « human ranges to 1-based well lists » │
# └────────────────────────────────────────────────────────────┘


def parse_well_range(text: str) -> list[int]:
    """Parse a human well range into a sorted list of 1-based well numbers.

    Accepts the sheet's many dialects: ``"1 to 24"``, ``"1-24"``,
    ``"1,2 4-24"``, ``"1 to 19, 21, 22"``.  Non-numeric tokens (dose units
    such as ``10mM``) are ignored, so gradient text yields its occupied
    wells; flag gradients separately via :func:`is_gradient_text`.
    """
    s = str(text or "").strip().lower()
    if not s:
        return []
    s = re.sub(r"\s+", " ", s)              # collapse the sheet's double spaces
    s = s.replace(" to ", "-")
    wells: set[int] = set()
    for tok in re.split(r"[,\s]+", s):
        if not tok:
            continue
        m = re.fullmatch(r"(\d+)-(\d+)", tok)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            wells.update(range(min(a, b), max(a, b) + 1))
        elif tok.isdigit():
            wells.add(int(tok))
    return sorted(w for w in wells if 1 <= w <= N_WELLS)


def compact_wells(wells: list[int]) -> str:
    """Inverse of :func:`parse_well_range`: ``[1,2,4,..,24] -> "1,2,4-24"``."""
    ws = sorted(set(wells))
    out: list[str] = []
    i = 0
    while i < len(ws):
        j = i
        while j + 1 < len(ws) and ws[j + 1] == ws[j] + 1:
            j += 1
        out.append(str(ws[i]) if i == j else f"{ws[i]}-{ws[j]}")
        i = j + 1
    return ",".join(out)


def is_gradient_text(text: str) -> bool:
    """True if arena text names a drug dose (per-well gradient, not a group)."""
    return bool(_GRADIENT_RE.search(str(text or "")))


def normalise_transgene(target_gene: str, sgrna: str) -> str:
    """Build a transgene label from target gene + guide, dropping null tokens.

    ``("NeuroD2", "g20") -> "NeuroD2 g20"``;
    ``("no_xenopus_target", "g15_5MM") -> "no_xenopus_target g15_5MM"``;
    ``("null", "null") -> "WT"``.
    """
    parts = [
        str(x).strip()
        for x in (target_gene, sgrna)
        if str(x or "").strip().lower() not in ("", "null", "none", "n/a", "nan")
    ]
    return " ".join(parts) if parts else "WT"


def parse_sheet_date(raw: object) -> datetime | None:
    """Parse a fertilisation/experiment date; return None if impossible.

    openpyxl hands real dates back as ``datetime``; only the messy string
    cases (e.g. the impossible ``"31/11/24"``) fall through to None.
    """
    if isinstance(raw, datetime):
        return raw
    s = str(raw or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y",
                "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ┌────────────────────────────────────────────────────────────┐
# │ Spreadsheet parsing  « two-row header to PlateRecord rows » │
# └────────────────────────────────────────────────────────────┘


@dataclass
class PlateRecord:
    """One spreadsheet row: a plate of one group across its occupied wells."""

    key: str
    initials: str
    female_identifier: str
    background_strain: str
    female_tank: str
    target_gene: str
    sgrna: str
    seq_folder: str
    fert_date_raw: str
    fert_date: datetime | None
    exp_date: datetime | None
    control_arenas: str
    test_arenas: str
    drug_dose: str
    notes: str
    videos: list[str] = field(default_factory=list)

    @property
    def transgene(self) -> str:
        return normalise_transgene(self.target_gene, self.sgrna)

    @property
    def arena_role(self) -> str:
        return "control" if parse_well_range(self.control_arenas) else "test"

    @property
    def is_gradient(self) -> bool:
        return is_gradient_text(self.test_arenas) or is_gradient_text(self.drug_dose)

    @property
    def occupied_wells(self) -> list[int]:
        ctrl = parse_well_range(self.control_arenas)
        return ctrl if ctrl else parse_well_range(self.test_arenas)


def _header_index_map(header_row: tuple) -> dict[str, int]:
    """Map our canonical field names to column indices via substring match."""
    idx: dict[str, int] = {}
    for col, cell in enumerate(header_row):
        label = str(cell or "").strip().lower()
        if not label:
            continue
        for needle, name in _FIELD_KEYS.items():
            if needle in label and name not in idx:
                idx[name] = col
    return idx


def parse_metadata_xlsx(xlsx_path: Path) -> list[PlateRecord]:
    """Parse the two-row-header metadata sheet into PlateRecord rows."""
    try:
        import openpyxl
    except ImportError as exc:                      # pragma: no cover
        raise ImportError(
            "metadata_ingest needs openpyxl: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    idx = _header_index_map(rows[1])

    def cell(row: tuple, name: str) -> str:
        col = idx.get(name)
        if col is None or col >= len(row):
            return ""
        val = row[col]
        return "" if val is None else str(val).strip()

    video_cols = [c for c in range(len(rows[1]))
                  if str(rows[1][c] or "").strip().lower().startswith("video")]

    records: list[PlateRecord] = []
    for row in rows[2:]:
        if idx.get("key") is None or not str(row[idx["key"]] or "").strip():
            continue
        fert_raw = row[idx["fert_date"]] if idx.get("fert_date") is not None else ""
        videos = [str(row[c]).strip() for c in video_cols
                  if c < len(row) and row[c] and str(row[c]).strip()]
        records.append(PlateRecord(
            key=cell(row, "key"),
            initials=cell(row, "initials"),
            female_identifier=cell(row, "female_identifier"),
            background_strain=cell(row, "background_strain"),
            female_tank=cell(row, "female_tank"),
            target_gene=cell(row, "target_gene"),
            sgrna=cell(row, "sgrna"),
            seq_folder=cell(row, "seq_folder"),
            fert_date_raw="" if fert_raw is None else str(fert_raw).strip(),
            fert_date=parse_sheet_date(fert_raw),
            exp_date=parse_sheet_date(row[idx["exp_date"]]
                                     if idx.get("exp_date") is not None else ""),
            control_arenas=cell(row, "control_arenas"),
            test_arenas=cell(row, "test_arenas"),
            drug_dose=cell(row, "drug_dose"),
            notes=cell(row, "notes"),
            videos=videos,
        ))
    return records


# ┌────────────────────────────────────────────────────────────┐
# │ Video matching  « filename to sheet row, three tiers »     │
# └────────────────────────────────────────────────────────────┘


def _norm(name: str) -> str:
    """Lowercase, drop non-alphanumerics, strip a trailing ``mp4`` extension.

    Stripping ``mp4`` *after* removing separators absorbs the sheet's
    ``,mp4`` comma-for-dot typos as well as the normal ``.mp4``.
    """
    s = re.sub(r"[^a-z0-9]", "", str(name or "").lower())
    return re.sub(r"mp4$", "", s)


def candidate_keys(video_name: str, records: list[PlateRecord]) -> list[str]:
    """Row keys whose any video matches by exact, tail, or prefix tier.

    Used to make an unmatched/ambiguous plan row actionable: it names the
    sheet rows the user should choose between.
    """
    nf = _norm(video_name)
    tf = re.search(r"(\d{3}\d{6})$", nf)
    pf = re.match(r"^(.*?\d{3})\d{6}$", nf)
    keys: set[str] = set()
    for rec in records:
        for vid in rec.videos:
            nv = _norm(vid)
            tv = re.search(r"(\d{3}\d{6})$", nv)
            pv = re.match(r"^(.*?\d{3})\d{6}$", nv)
            if nv == nf or (tf and tv and tf.group(1) == tv.group(1)) \
                    or (pf and pv and pf.group(1) == pv.group(1)):
                keys.add(rec.key)
    return sorted(keys)


def match_video(video_name: str, records: list[PlateRecord]) -> tuple[PlateRecord | None, str]:
    """Match one video filename to its PlateRecord; return (record, method).

    Three tiers, tried in order and required to converge to a single row:
    ``exact`` normalised filename, ``tail`` (index + HHMMSS timestamp), and
    ``prefix`` (filename minus the trailing 6-digit time).  Returns
    (None, "unmatched"/"ambiguous") when no single row wins.
    """
    nf = _norm(video_name)

    by_exact: dict[str, list[PlateRecord]] = {}
    by_tail: dict[str, list[PlateRecord]] = {}
    by_prefix: dict[str, list[PlateRecord]] = {}
    for rec in records:
        for vid in rec.videos:
            nv = _norm(vid)
            by_exact.setdefault(nv, []).append(rec)
            mt = re.search(r"(\d{3}\d{6})$", nv)
            if mt:
                by_tail.setdefault(mt.group(1), []).append(rec)
            mp = re.match(r"^(.*?\d{3})\d{6}$", nv)
            if mp:
                by_prefix.setdefault(mp.group(1), []).append(rec)

    if nf in by_exact and len({r.key for r in by_exact[nf]}) == 1:
        return by_exact[nf][0], "exact"
    mt = re.search(r"(\d{3}\d{6})$", nf)
    if mt and len({r.key for r in by_tail.get(mt.group(1), [])}) == 1:
        return by_tail[mt.group(1)][0], "tail"
    mp = re.match(r"^(.*?\d{3})\d{6}$", nf)
    if mp and len({r.key for r in by_prefix.get(mp.group(1), [])}) == 1:
        return by_prefix[mp.group(1)][0], "prefix"

    hits = by_exact.get(nf, [])
    return (None, "ambiguous" if hits else "unmatched")


def list_videos(folder: Path) -> list[Path]:
    """Recursively list .mp4 files, skipping hidden ._ AppleDouble stubs."""
    out = [p for p in sorted(Path(folder).rglob("*.mp4"))
           if not p.name.startswith(".")]
    return out


# ┌────────────────────────────────────────────────────────────┐
# │ Database get-or-create  « idempotent Frog / group / well type » │
# └────────────────────────────────────────────────────────────┘


def get_or_create_frog(db: DatabaseHandler, female_identifier: str,
                       background_strain: str, female_tank: str) -> Frog:
    """Fetch the matching Frog (mother) or create it."""
    found = db.get_records(Frog, filters={
        "female_identifier": female_identifier,
        "background_strain": background_strain,
    })
    if found:
        return found[0]
    tank = female_tank if str(female_tank or "").strip() else None
    frog = Frog(female_tank=tank, female_identifier=female_identifier,
                background_strain=background_strain)
    db.add_record(frog)
    return frog


def get_or_create_tadpole_group(db: DatabaseHandler, mother_id: int,
                                fert_date: datetime | None, transgene: str,
                                seq_folder: str) -> TadpoleGroup:
    """Fetch the matching TadpoleGroup or create it (keyed on mother+transgene)."""
    found = db.get_records(TadpoleGroup, filters={
        "mother_id": mother_id,
        "transgene": transgene,
    })
    for grp in found:
        if grp.fertilisation_date == fert_date:
            return grp
    grp = TadpoleGroup(mother_id=mother_id, fertilisation_date=fert_date,
                       development_stage=None, seq_folder=seq_folder or None,
                       transgene=transgene)
    db.add_record(grp)
    return grp


def get_or_create_well_type(db: DatabaseHandler, name: str = DEFAULT_WELL_TYPE_NAME,
                            description: str = DEFAULT_WELL_TYPE_DESC) -> WellType:
    """Fetch the named WellType or create it."""
    found = db.get_records(WellType, filters={"name": name})
    if found:
        return found[0]
    wt = WellType(name=name, description=description)
    db.add_record(wt)
    return wt


def get_or_create_investigator(db: DatabaseHandler, initials: str) -> Investigator:
    """Fetch an Investigator whose name initials match, else create from initials."""
    initials = str(initials or "").strip()
    for inv in db.get_records(Investigator):
        first = str(inv.first_name or "").strip()
        last = str(inv.last_name or "").strip()
        sig = (first[:1] + last[:1]).upper()
        if first.upper() == initials.upper() or (sig and sig == initials.upper()[:2]):
            return inv
    inv = Investigator(first_name=initials, last_name="")
    db.add_record(inv)
    return inv


# ┌────────────────────────────────────────────────────────────┐
# │ Plate layout + CSV  « 24-well table the pipeline ingests » │
# └────────────────────────────────────────────────────────────┘


def build_plate_layout(occupied_wells: list[int], group_id: int,
                       well_type_id: int, n_wells: int = N_WELLS) -> dict:
    """Build the per-well layout dict PlateManager.manage_plate() would return.

    Occupied wells get the group + well-type id; the rest stay None and are
    dropped from grouped analyses at ingest.
    """
    occ = set(occupied_wells)
    well_type_ids = [well_type_id if (w + 1) in occ else None for w in range(n_wells)]
    tadpole_type_ids = [group_id if (w + 1) in occ else None for w in range(n_wells)]
    return {"well_type_ids": well_type_ids, "tadpole_type_ids": tadpole_type_ids}


def write_meta_data_csv(out_csv: Path, layout: dict, investigator_id: int,
                        experiment_type_id: int) -> None:
    """Write the 24-row per-well metadata CSV result_manager ingests."""
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    n = len(layout["well_type_ids"])
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["well_number", "investigator_id", "experiment_type_id",
                    "well_type_ids", "tadpole_type_ids"])
        for i in range(n):
            wt = layout["well_type_ids"][i]
            tg = layout["tadpole_type_ids"][i]
            w.writerow([i, investigator_id, experiment_type_id,
                        "" if wt is None else wt, "" if tg is None else tg])


# ┌────────────────────────────────────────────────────────────┐
# │ Plan CSV round-trip  « reviewable map between sheet and runs » │
# └────────────────────────────────────────────────────────────┘


_PLAN_FIELDS = [
    "video_path", "video_stem", "row_key", "match_method",
    "investigator_initials", "female_identifier", "background_strain",
    "female_tank", "target_gene", "sgrna", "transgene",
    "fert_date", "fert_date_raw", "exp_date", "seq_folder",
    "occupied_wells", "n_occupied", "is_gradient", "arena_role", "notes",
]


def build_plan(video_paths: list[Path], records: list[PlateRecord]) -> list[dict]:
    """Match each video to a record and produce one reviewable plan row each."""
    plan: list[dict] = []
    for vp in video_paths:
        rec, method = match_video(vp.name, records)
        if rec is None:
            cands = candidate_keys(vp.name, records)
            note = "NEEDS MANUAL MATCH"
            if cands:
                note += "; candidates: " + ", ".join(cands)
            plan.append({"video_path": str(vp), "video_stem": vp.stem,
                         "row_key": "", "match_method": method,
                         "occupied_wells": "", "n_occupied": 0,
                         "is_gradient": "", "notes": note})
            continue
        wells = rec.occupied_wells
        plan.append({
            "video_path": str(vp), "video_stem": vp.stem,
            "row_key": rec.key, "match_method": method,
            "investigator_initials": rec.initials,
            "female_identifier": rec.female_identifier,
            "background_strain": rec.background_strain,
            "female_tank": rec.female_tank,
            "target_gene": rec.target_gene, "sgrna": rec.sgrna,
            "transgene": rec.transgene,
            "fert_date": rec.fert_date.date().isoformat() if rec.fert_date else "",
            "fert_date_raw": rec.fert_date_raw,
            "exp_date": rec.exp_date.date().isoformat() if rec.exp_date else "",
            "seq_folder": rec.seq_folder,
            "occupied_wells": compact_wells(wells), "n_occupied": len(wells),
            "is_gradient": "1" if rec.is_gradient else "0",
            "arena_role": rec.arena_role, "notes": rec.notes,
        })
    return plan


def write_plan_csv(out_csv: Path, plan: list[dict]) -> None:
    """Write the plan rows to a CSV the user can review and edit before commit."""
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=_PLAN_FIELDS)
        w.writeheader()
        for row in plan:
            w.writerow({k: row.get(k, "") for k in _PLAN_FIELDS})


def read_plan_csv(in_csv: Path) -> list[dict]:
    """Read a (possibly hand-edited) plan CSV back into dict rows."""
    with open(Path(in_csv), newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


# ┌────────────────────────────────────────────────────────────┐
# │ CLI  « python -m tadpose.metadata_ingest plan|commit »     │
# └────────────────────────────────────────────────────────────┘


def cmd_plan(args: argparse.Namespace) -> None:
    """plan: parse the sheet, match videos, write the reviewable plan CSV."""
    records = parse_metadata_xlsx(args.xlsx)
    videos = list_videos(args.videos)
    plan = build_plan(videos, records)
    out = args.out or (Path(args.videos) / "video_group_plan.csv")
    write_plan_csv(out, plan)

    matched = [r for r in plan if r["row_key"]]
    unmatched = [r for r in plan if not r["row_key"]]
    grad = [r for r in matched if str(r.get("is_gradient")) == "1"]
    print(f"parsed {len(records)} sheet rows; matched {len(matched)}/{len(videos)} videos")
    for r in matched:
        print(f"  {r['video_stem']:<32} -> {r['transgene']:<26} "
              f"clutch {r['female_identifier']:<3} wells {r['occupied_wells']:<10} "
              f"[{r['match_method']}]")
    for r in unmatched:
        print(f"  {r['video_stem']:<32} -> UNMATCHED ({r['match_method']})")
    if grad:
        print(f"\n  {len(grad)} matched video(s) look like drug GRADIENT plates; "
              "their per-well well-types need manual handling:")
        for r in grad:
            print(f"    {r['video_stem']}  ({r['notes']})")
    if unmatched:
        print(f"\n  {len(unmatched)} unmatched: edit row_key/transgene/occupied_wells "
              "in the plan CSV before commit.")
    print(f"\nwrote plan: {out}")
    print("Review/edit it, then: python -m tadpose.metadata_ingest commit "
          f"--plan {out} --db <db> --output <out>")


def cmd_commit(args: argparse.Namespace) -> None:
    """commit: seed the DB from the plan and write per-video metadata CSVs."""
    plan = read_plan_csv(args.plan)
    db_handler = DatabaseHandler(f"sqlite:///{args.db}")
    written: list[Path] = []
    skipped: list[str] = []

    with db_handler as db:
        well_type = get_or_create_well_type(db)
        for row in plan:
            if not row.get("transgene") or not row.get("video_path"):
                skipped.append(row.get("video_stem", "?") + " (no transgene/path)")
                continue
            if str(row.get("is_gradient")) == "1" and not args.allow_gradient:
                skipped.append(row.get("video_stem", "?") + " (gradient; --allow-gradient)")
                continue
            wells = parse_well_range(row.get("occupied_wells", ""))
            if not wells:
                skipped.append(row.get("video_stem", "?") + " (no occupied wells)")
                continue

            frog = get_or_create_frog(db, row["female_identifier"],
                                      row["background_strain"], row.get("female_tank", ""))
            fert = parse_sheet_date(row.get("fert_date") or args.default_fert_date or "")
            group = get_or_create_tadpole_group(db, frog.frog_id, fert,
                                                row["transgene"], row.get("seq_folder", ""))
            investigator = get_or_create_investigator(db, row.get("investigator_initials", ""))

            layout = build_plate_layout(wells, group.tadpole_group_id, well_type.well_type_id)
            out_csv = Path(args.output) / row["video_stem"] / "meta_data" / "meta_data_table.csv"
            write_meta_data_csv(out_csv, layout, investigator.investigator_id,
                                args.experiment_type_id)
            written.append(out_csv)
            print(f"  {row['video_stem']:<32} group {group.tadpole_group_id} "
                  f"({row['transgene']}) wells {compact_wells(wells)} -> {out_csv}")

    print(f"\nwrote {len(written)} metadata CSV(s); seeded groups in {args.db}")
    if skipped:
        print(f"skipped {len(skipped)}:")
        for s in skipped:
            print(f"  - {s}")


def build_arg_parser() -> argparse.ArgumentParser:
    # Defaults resolve from the active profile (videos_root / db_path /
    # output_root), falling back to data_root-relative paths, so --db / --output
    # / --videos need not be typed.
    videos_default = config.configured_path("videos_root", "videos", "raw", "nov2024_genes")
    db_default = config.configured_path("db_path", "databases", "xenopus_DEE.sqlite3")
    output_default = config.configured_path("output_root", "results", "pipeline", "nov2024_new_genes")

    p = argparse.ArgumentParser(
        prog="python -m tadpose.metadata_ingest",
        description="Spreadsheet -> plate groups + per-video metadata CSVs.")
    sub = p.add_subparsers(dest="command", required=True)

    pp = sub.add_parser("plan", help="parse sheet, match videos, write plan CSV")
    pp.add_argument("--xlsx", type=Path, default=videos_default / "METADATA_TADPOLES.xlsx",
                    help="metadata spreadsheet (default: <videos>/METADATA_TADPOLES.xlsx)")
    pp.add_argument("--videos", type=Path, default=videos_default, help="raw-video folder")
    pp.add_argument("--out", type=Path, default=None, help="plan CSV path")
    pp.set_defaults(func=cmd_plan)

    pc = sub.add_parser("commit", help="seed DB + write per-video metadata CSVs")
    pc.add_argument("--plan", type=Path, required=True, help="reviewed plan CSV")
    pc.add_argument("--db", type=Path, default=db_default, help="SQLite database")
    pc.add_argument("--output", type=Path, default=output_default, help="per-video output base")
    pc.add_argument("--experiment-type-id", type=int, default=DEFAULT_EXPERIMENT_TYPE_ID)
    pc.add_argument("--default-fert-date", type=str, default=None,
                    help="fallback fertilisation date (YYYY-MM-DD) for unparseable rows")
    pc.add_argument("--allow-gradient", action="store_true",
                    help="also write gradient plates (uniform well type; doses lost)")
    pc.set_defaults(func=cmd_commit)
    return p


def main(argv: list[str] | None = None) -> None:
    args = build_arg_parser().parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
